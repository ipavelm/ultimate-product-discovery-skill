#!/usr/bin/env python3
"""Expand the formulas from column C (month 1) into D-N (months 2-12) on P&L and Cash Flow.

The financial-plan-template.xlsx template carries formulas only in column C.
Columns D..N (months 2..12) are empty. Without expanding them the plan shows
revenue for month 1 alone, which is useless to an investor.

The script walks the P&L and Cash Flow sheets, finds the rows holding formulas in
C and expands them into D..N, shifting the relative references. Absolute
references ($X) are left alone.

Usage:
    python3 scripts/roll_formulas.py /home/claude/financial-plan.xlsx

Afterwards you must recalculate:
    python3 /home/claude/scripts/recalc.py /home/claude/financial-plan.xlsx
"""
import sys
import re

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("❌ openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def shift_formula(formula: str, delta: int):
    """Shift the relative column references by delta.

    Absolute references ($C$5) are left alone. It handles these forms:
    - A1, AB5 — relative, shifted
    - $A1, $AB5 — column-absolute, not shifted
    - A$1 — column-relative, shifted (the row is pinned)
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return None

    def shift_ref(m):
        full = m.group(0)
        # Check whether there is a $ before the column letters
        if full.startswith('$'):
            return full  # column-absolute — leave it
        col_letters = m.group(1)
        row_num = m.group(2)
        col_idx = column_index_from_string(col_letters)
        new_col_idx = col_idx + delta
        if new_col_idx < 1:
            return full
        return f"{get_column_letter(new_col_idx)}{row_num}"

    # Reference regex: optional $, column letters, row digits
    return re.sub(r'\$?([A-Z]+)(\d+)', shift_ref, formula)


def roll_sheet(ws, start_col=4, end_col=15):
    """Expand the formulas from column C into columns [start_col..end_col).

    The default is D=4..N=14 inclusive (end_col=15 is exclusive).

    Returns the tuple (rolled, skipped): how many were expanded and how many were
    skipped because the cell already held a value.
    """
    rolled = 0
    skipped = 0
    for row in range(1, ws.max_row + 1):
        c_value = ws.cell(row=row, column=3).value
        if not isinstance(c_value, str) or not c_value.startswith("="):
            continue
        for col in range(start_col, end_col):
            if ws.cell(row=row, column=col).value is not None:
                skipped += 1
                continue
            shifted = shift_formula(c_value, col - 3)
            if shifted:
                ws.cell(row=row, column=col).value = shifted
                rolled += 1
    return rolled, skipped


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path = sys.argv[1]

    try:
        wb = openpyxl.load_workbook(path)
    except FileNotFoundError:
        print(f"❌ File not found: {path}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Could not open the file: {e}")
        sys.exit(1)

    total_rolled = 0
    for sheet_name in ["P&L", "Cash Flow"]:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Sheet \"{sheet_name}\" not found — skipping")
            continue
        rolled, skipped = roll_sheet(wb[sheet_name])
        total_rolled += rolled
        print(f"✅ {sheet_name}: {rolled} formulas expanded"
              + (f" ({skipped} already-filled cells skipped)" if skipped else ""))

    if total_rolled == 0:
        print()
        print("⚠️  Nothing was expanded — the formulas may already be in place,")
        print("   or the file does not match the template's expected structure.")
        sys.exit(0)

    wb.save(path)
    print()
    print(f"✅ Saved: {path}")
    print()
    print("Next step — recalculate the values through LibreOffice:")
    print(f"    python3 /home/claude/scripts/recalc.py {path}")
    print()
    print("Expected result: \"status\": \"success\", \"total_errors\": 0")


if __name__ == "__main__":
    main()
