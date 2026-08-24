#!/usr/bin/env python3
"""Check that the skill itself is internally consistent.

This is not the pre-flight check — `preflight_check.sh` verifies the environment
before a PD run. This one verifies the skill's own files after they are edited:
frontmatter, internal links, sheet names the instructions point at, formula
integrity, leftover third-party data, and placeholder coverage in the templates.

The last step recalculates the financial-plan template and asserts that the
model still produces possible numbers — no negative customers, no negative
revenue, a descending funnel. That is the class of defect that is invisible in
the formula text: before it was fixed, the template returned -1 customers and
-9,385 revenue in month 2. It needs the `formulas` package and takes several
minutes on this workbook; it is skipped with a note when the package is missing,
and `--no-recalc` skips it when you only want the structural checks.

Usage:
    python3 scripts/self_check.py             # run from the skill directory
    python3 scripts/self_check.py <path>      # or point it at one
    python3 scripts/self_check.py --no-recalc # structure only, skip the recalculation

Exit code 0 when everything passes, 1 when any check fails.
"""
import re
import sys
import zipfile
from pathlib import Path

args = [a for a in sys.argv[1:] if not a.startswith('--')]
RECALC = '--no-recalc' not in sys.argv
ROOT = Path(args[0] if args else Path(__file__).resolve().parent.parent)
CYRILLIC = re.compile('[\u0400-\u052f]')  # written as escapes so this file stays ASCII
# Cyrillic that is expected: the localised name of Excel's built-in Normal style,
# which Excel identifies by builtinId rather than by name.
ALLOWED_CYRILLIC = {'assets/financial-plan-template.xlsx': 7}

failures = []
notes = []


def check(name, ok, detail=''):
    print(f"  {'✅' if ok else '❌'} {name}" + (f" — {detail}" if detail else ''))
    if not ok:
        failures.append(name)


def decode_refs(raw: bytes) -> str:
    """Office parts store non-ASCII as numeric character references; expand them."""
    text = raw.decode('utf-8', 'replace')
    return re.sub(r'&#(\d+);',
                  lambda m: chr(int(m.group(1))) if int(m.group(1)) > 127 else m.group(0),
                  text)


def office_text(path: Path) -> str:
    z = zipfile.ZipFile(path)
    return ''.join(decode_refs(z.read(n)) for n in z.namelist() if n.endswith('.xml'))


print('Frontmatter')
try:
    import yaml
    body = (ROOT / 'SKILL.md').read_text(encoding='utf-8')
    m = re.match(r'^---\n(.*?)\n---\n', body, re.S)
    fm = yaml.safe_load(m.group(1)) if m else None
    check('parses as YAML into a mapping', isinstance(fm, dict))
    if isinstance(fm, dict):
        name = fm.get('name', '')
        desc = fm.get('description', '')
        check('name matches [a-z0-9-]{1,64}', bool(re.fullmatch(r'[a-z0-9-]{1,64}', name)), name)
        check('description within 1024 chars', len(desc) <= 1024, f'{len(desc)} chars')
        version = str(fm.get('metadata', {}).get('version', ''))
        stamped = re.search(r'SKILL_VERSION = "([^"]+)"',
                            (ROOT / 'scripts/init_kb.py').read_text(encoding='utf-8'))
        check('init_kb.py stamps the same version',
              bool(stamped) and stamped.group(1) == version,
              f'SKILL.md {version} vs init_kb {stamped.group(1) if stamped else "?"}')
except ImportError:
    notes.append('PyYAML missing — frontmatter not parsed (pip install pyyaml)')

print('\nInternal links')
broken = []
for md in [ROOT / 'SKILL.md', ROOT / 'README.md'] + sorted((ROOT / 'references').glob('*.md')):
    for link in re.findall(r'\]\(([^)#][^)]*)\)', md.read_text(encoding='utf-8')):
        if link.startswith(('http://', 'https://')):
            continue
        if not (md.parent / link).resolve().exists():
            broken.append(f'{md.name} → {link}')
check('every relative link resolves', not broken, '; '.join(broken[:5]))

print('\nScripts and assets the instructions name')
docs = '\n'.join(p.read_text(encoding='utf-8') for p in
                 [ROOT / 'SKILL.md'] + sorted((ROOT / 'references').glob('*.md')))
# Only paths relative to the skill; an absolute path such as
# /home/claude/scripts/recalc.py points at a different skill's copy.
refs = sorted(set(re.findall(r'(?<![\w/])(?:scripts|assets)/[\w.-]+\.(?:py|sh|pptx|xlsx|docx)', docs)))
missing = [ref for ref in refs if not (ROOT / ref).exists()]
check('all referenced files exist', not missing, ', '.join(missing) or f'{len(refs)} checked')

print('\nWorkbook')
wb_path = ROOT / 'assets/financial-plan-template.xlsx'
try:
    import openpyxl
    wb = openpyxl.load_workbook(wb_path)
    formulas = [c.value for ws in wb.worksheets for row in ws.iter_rows()
                for c in row if isinstance(c.value, str) and c.value.startswith('=')]
    check('opens and has sheets', len(wb.sheetnames) > 0, f'{len(wb.sheetnames)} sheets')
    check('no error values in formulas',
          not [f for f in formulas if re.search(r'#(REF|NAME\?|VALUE|DIV/0)', f)],
          f'{len(formulas)} formulas')
    sheet_refs = set()
    for f in formulas:
        sheet_refs |= set(re.findall(r"'([^']+)'!", f))
    dangling = sorted(sheet_refs - set(wb.sheetnames))
    check('every cross-sheet reference resolves', not dangling, ', '.join(dangling))
    named = set(re.findall(r"='([^']+)'!", docs))
    unknown = sorted(named - set(wb.sheetnames))
    check('sheet names used in the docs exist', not unknown, ', '.join(unknown))
except ImportError:
    notes.append('openpyxl missing — workbook not checked (pip install openpyxl)')

print('\nTemplates')
for rel, tag in [('assets/presentation-template.pptx', 'a:t'),
                 ('assets/one-pager-template.pptx', 'a:t'),
                 ('assets/interview-guide-template.docx', 'w:t')]:
    path = ROOT / rel
    if not path.exists():
        check(f'{Path(rel).name} present', False)
        continue
    text = office_text(path)
    runs = [s for s in re.findall(rf'<{tag}(?:\s[^>]*)?>([^<]*)</{tag}>', text) if s.strip()]
    placeholders = [s for s in runs if re.search(r'\[[^\]]{1,80}\]|\{\{[^}]+\}\}', s)]
    share = 100 * len(placeholders) // max(len(runs), 1)
    check(f'{Path(rel).name} keeps placeholders', share >= 10,
          f'{len(placeholders)}/{len(runs)} runs ({share}%)')

print('\nLeftover data')
for path in sorted(p for p in ROOT.rglob('*') if p.is_file()
                   and '.git' not in p.parts and '__pycache__' not in p.parts):
    rel = str(path.relative_to(ROOT))
    if path.suffix in ('.xlsx', '.pptx', '.docx'):
        count = len(CYRILLIC.findall(office_text(path)))
    elif path.suffix in ('.md', '.py', '.sh', '.json'):
        count = len(CYRILLIC.findall(path.read_text(encoding='utf-8')))
    else:
        continue
    check(f'{rel}: no unexpected Cyrillic', count <= ALLOWED_CYRILLIC.get(rel, 0),
          f'{count} chars' if count else '')

if RECALC:
    print('\nRecalculation (financial-plan template)')
    try:
        import contextlib
        import io
        import shutil
        import tempfile
        import warnings

        import formulas

        warnings.filterwarnings('ignore')
        with tempfile.TemporaryDirectory() as tmp:
            # formulas keys cells by the uppercased file name, so copy to a known one
            work = Path(tmp) / 'MODEL.xlsx'
            shutil.copy(wb_path, work)
            hush = io.StringIO()
            with contextlib.redirect_stderr(hush), contextlib.redirect_stdout(hush):
                solution = formulas.ExcelModel().loads(str(work)).finish().calculate()

        cells = {k.upper(): v for k, v in solution.items()}

        def value(sheet, ref):
            cell = cells.get(f"'[MODEL.XLSX]{sheet}'!{ref}".upper())
            if cell is None:
                return None
            raw = getattr(cell, 'value', cell)
            try:
                raw = raw[0, 0]
            except Exception:
                pass
            try:
                return float(raw)
            except Exception:
                return None

        months = [chr(c) for c in range(ord('C'), ord('N') + 1)]   # months 1-12

        def series(row):
            return [(col, value('MODEL', f'{col}{row}')) for col in months]

        check('the workbook recalculates', bool(cells), f'{len(cells)} cells')

        # A plan can hold any assumptions, but never a negative headcount or negative revenue.
        COUNTS_AND_MONEY = {28: 'total paying customers', 31: 'new paying (segment 2)',
                            32: 'new paying (segment 1)', 33: 'returning (segment 2)',
                            34: 'returning (segment 1)', 41: 'total GMV',
                            42: 'GMV (segment 2)', 43: 'GMV (segment 1)'}
        negatives = [f'{label} {col}: {val:,.0f}'
                     for row, label in COUNTS_AND_MONEY.items()
                     for col, val in series(row)
                     if val is not None and val < -0.5]
        check('no negative customer counts or revenue', not negatives, '; '.join(negatives[:4]))

        funnel = [value('MODEL', f'C{row}') for row in (18, 20, 22, 24, 26)]
        check('month-1 funnel is positive',
              all(v is not None and v > 0 for v in funnel),
              ' → '.join('?' if v is None else f'{v:,.0f}' for v in funnel))
        check('month-1 funnel descends',
              all(a is not None and b is not None and a >= b
                  for a, b in zip(funnel, funnel[1:])))

        gmv = [v for _, v in series(41) if v is not None]
        check('total GMV is positive every month', bool(gmv) and min(gmv) > 0,
              f'month 1 {gmv[0]:,.0f} → month {len(gmv)} {gmv[-1]:,.0f}' if gmv else 'no values')
        check('total GMV does not shrink',
              all(a <= b + 1e-6 for a, b in zip(gmv, gmv[1:])))

        paying = [v for _, v in series(28) if v is not None]
        check('paying customers do not shrink',
              all(a <= b + 1e-6 for a, b in zip(paying, paying[1:])),
              f'{paying[0]:,.0f} → {paying[-1]:,.0f}' if paying else 'no values')

        profit = [v for _, v in series(60) if v is not None]
        check('operating profit is computable every month', len(profit) == len(months),
              f'{len(profit)}/{len(months)} months')

        # Sign checks alone are too weak: a formula can point at the wrong row and
        # still produce a positive, rising number. These compare rows that are
        # computed independently of each other.
        # A segment's revenue cannot be less than its new customers times their
        # price — returning customers can only add to it.
        floor_breaks = []
        for gmv_row, new_row, price_row, label in ((42, 31, 35, 'segment 2'),
                                                   (43, 32, 37, 'segment 1')):
            for col in months:
                gmv = value('MODEL', f'{col}{gmv_row}')
                new = value('MODEL', f'{col}{new_row}')
                price = value('MODEL', f'{col}{price_row}')
                if None in (gmv, new, price):
                    continue
                if gmv + 1 < new * price:
                    floor_breaks.append(f'{label} {col}: {gmv:,.0f} < {new * price:,.0f}')
        check('segment revenue covers new customers x their price', not floor_breaks,
              '; '.join(floor_breaks[:3]))

        # Total paying customers must be the sum of its four component rows.
        sum_breaks = []
        for col in months:
            total = value('MODEL', f'{col}28')
            parts = [value('MODEL', f'{col}{r}') for r in (31, 32, 33, 34)]
            if total is None or None in parts:
                continue
            if abs(total - sum(parts)) > 0.5:
                sum_breaks.append(f'{col}: {total:,.1f} vs {sum(parts):,.1f}')
        check('total paying customers equals its parts', not sum_breaks,
              '; '.join(sum_breaks[:3]))

        # New paying customers are the last funnel stage, split between segments.
        funnel_breaks = []
        for col in months:
            stage = value('MODEL', f'{col}26')
            split = [value('MODEL', f'{col}31'), value('MODEL', f'{col}32')]
            if stage is None or None in split:
                continue
            if abs(stage - sum(split)) > 0.5:
                funnel_breaks.append(f'{col}: funnel {stage:,.1f} vs new {sum(split):,.1f}')
        check('new paying customers match the funnel output', not funnel_breaks,
              '; '.join(funnel_breaks[:3]))
    except ImportError:
        notes.append('formulas missing — the model was not recalculated '
                     '(pip install formulas), rerun with --no-recalc to silence this')
    except Exception as exc:
        check('the workbook recalculates', False, f'{type(exc).__name__}: {exc}')
else:
    notes.append('recalculation skipped (--no-recalc)')

print()
for note in notes:
    print(f'  ⚠️  {note}')
if failures:
    print(f'\n❌ {len(failures)} check(s) failed:')
    for f in failures:
        print(f'   - {f}')
    sys.exit(1)
print('\n✅ All checks passed.')
