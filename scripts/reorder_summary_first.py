#!/usr/bin/env python3
"""Move the 'Summary' sheet to first position in the financial plan.

Used after the plan is filled in, so the investor or CEO sees the executive view
the moment they open the file instead of scrolling to the 14th sheet.

Usage:
    python3 reorder_summary_first.py <path-to-financial-plan.xlsx>

The file is modified in place.
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def reorder_summary_first(xlsx_path: str, summary_sheet_name: str = "Summary") -> bool:
    """Move the named sheet to first position. True on success, False if it is absent."""
    wb = load_workbook(xlsx_path)
    if summary_sheet_name not in wb.sheetnames:
        print(f"⚠️  Sheet '{summary_sheet_name}' not found in the file.")
        print(f"   Available sheets: {wb.sheetnames}")
        return False

    # openpyxl keeps the sheets in wb._sheets (a list) — move by index
    summary_sheet = wb[summary_sheet_name]
    current_index = wb._sheets.index(summary_sheet)
    if current_index == 0:
        print(f"✅ Sheet '{summary_sheet_name}' is already first, nothing to do.")
        return True

    # move_sheet with offset = -current_index moves it to the front
    wb.move_sheet(summary_sheet, offset=-current_index)
    wb.save(xlsx_path)
    print(f"✅ Sheet '{summary_sheet_name}' moved from position {current_index + 1} to position 1.")
    return True


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"❌ File not found: {path}")
        sys.exit(1)

    sheet_name = sys.argv[2] if len(sys.argv) == 3 else "Summary"
    ok = reorder_summary_first(str(path), sheet_name)
    sys.exit(0 if ok else 1)
